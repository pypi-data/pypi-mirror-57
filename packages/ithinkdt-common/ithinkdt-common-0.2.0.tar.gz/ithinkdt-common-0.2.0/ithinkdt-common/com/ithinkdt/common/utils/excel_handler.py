"""
excel操作类
主要用于读取Excel表格中的测试用例和数据，一个模块一个sheet，一个sheet包含该模块的所有冒烟用例
"""
from openpyxl import load_workbook


class ExcelHandler:
    """excel 封装"""
    # 测试数据，写代码之前之前已经写好测试数据的Excel. 文件名
    def __init__(self, file_name, sheet_name):
        """
        excel 封装,解析
        :param file_name: Excel表格的名称，路径拼接到.xlsx
        :param sheet_name: 表格中sheet名称
        """
        self.file_name = file_name
        self.sheet_name = sheet_name
        # self.wb = load_workbook()

    def open(self):
        self.wb = load_workbook(self.file_name)
        if isinstance(self.sheet_name, int):
            self.sheet = self.wb.worksheets[self.sheet_name]
        else:
            self.sheet = self.wb[self.sheet_name]

    def headers(self):
        """获取标题"""
        self.open()
        headers = [c.value for c in self.sheet[1]]
        self.wb.close()
        return headers

    def read(self, start_row=2, start_column=1):
        """获取所有的数据"""
        self.open()
        sheet = self.sheet
        header = [c.value for c in sheet[1]]
        data = []
        for row in range(start_row, sheet.max_row + 1):
            row_data = []
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            row_data = dict(zip(header, row_data))
            data.append(row_data)
        self.wb.close()
        return data

    def save(self):
        """保存"""
        self.wb.save(self.file_name)
        self.wb.close()

    def write(self, row, column, data):
        self.open()
        self.sheet.cell(row, column).value = data
        self.save()


# if __name__ == '__main__':
#     xls = ExcelHandler(r'D:\python36\project\srm_UITest\srm-ui-test\srm__webUI\srm_WebUi\data\testCase.xlsx', 'receiptGoods')
#     print(xls.read())